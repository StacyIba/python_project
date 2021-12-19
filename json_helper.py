import json as json_lib
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


class JsonHelper:
    def __init__(self,
                 input_json_file_name="DM.json",
                 output_json_file_name="result.json",
                 output_excel_file_name="result.xlsx",

                 root_calc_element_name="calculation",
                 root_filter_element_name="filter",
                 root_metadata_element_name="metadataTreeView",

                 root_folder_element_name="folderItem",
                 folder_element_name="folder",
                 ref_element_name="ref",
                 label_element_name="label",
                 identifier_field_name="identifier",
                 hidden_field_name="hidden",

                 domains_to_exclude=("Enterprise_Performance_Management",),

                 columns_to_output_xlsx=None,

                 delete_hidden=False):

        if columns_to_output_xlsx is None:
            columns_to_output_xlsx = {
                "A": ("Structure", "identifier", None),
                "B": ("Items / Filters", "label", None),
                "C": ("Description", "screenTip", None),
                "D": ("Hidden away", "hidden", "FALSE")
            }
        self.input_json_file_name = input_json_file_name
        self.output_json_file_name = output_json_file_name
        self.output_excel_file_name = output_excel_file_name

        self.root_calc_element_name = root_calc_element_name
        self.root_filter_element_name = root_filter_element_name
        self.root_metadata_element_name = root_metadata_element_name
        self.root_folder_element_name = root_folder_element_name
        self.folder_element_name = folder_element_name
        self.ref_element_name = ref_element_name
        self.label_element_name = label_element_name
        self.identifier_field_name = identifier_field_name

        self.domains_to_exclude = domains_to_exclude

        self.fields_to_select_from_calc_filter = [(field, value_by_default)
                                                  for key, (field_desc, field, value_by_default)
                                                  in columns_to_output_xlsx.items()]
        self.columns_to_output_xlsx = columns_to_output_xlsx

        if self.identifier_field_name not in self.columns_to_output_xlsx.get("A"):
            raise ValueError(f"identifier_field_name {self.identifier_field_name} must exist in columns_to_output_xlsx "
                             f"and be the first (A column)")

        self.hidden_field_name = hidden_field_name
        self.delete_hidden = delete_hidden
        if self.hidden_field_name:
            self.hidden_column_xlsx_list = [column_ident
                                            for column_ident, (field_desc, field, value_by_default) in
                                            columns_to_output_xlsx.items()
                                            if field == self.hidden_field_name]
        else:
            self.hidden_column_xlsx_list = []
        self.hidden_sheets = []

        self.folder_dict = {}
        self.folder_with_calc_filter = {}
        self.sheet_names = set()

        with open(self.input_json_file_name, encoding="utf-8") as f:
            self.json = json_lib.load(f)

    def write_to_json(self,
                      fields_list,
                      root_element_name,
                      input_json_file_name=None,
                      result_json_file_name=None,
                      ):
        if input_json_file_name:
            with open(input_json_file_name, encoding="utf-8") as f:
                json = json_lib.load(f)
        else:
            json = self.json

        if not result_json_file_name:
            result_json_file_name = self.output_json_file_name

        json_dict = {root_element_name: [JsonHelper.select_fields(
            fields=fields_list,
            dictionary=elem
        ) for elem in json[root_element_name]]}

        json.dump(json_dict, open(result_json_file_name, "w"), indent=2)

    def write_to_excel(self,
                       result_excel_file_name=None, ):
        if not result_excel_file_name:
            result_excel_file_name = self.output_excel_file_name

        self.get_folder_to_calc_and_filter()
        self.sheet_names = sorted(self.sheet_names)
        sheet_dict = {sheet_name: {"calc_filter_list": []} for sheet_name in self.sheet_names}
        for (folder, hidden), calc_filter in self.folder_with_calc_filter.items():
            main_folder, subfolder_label = folder.split("___")
            for sheet_name in self.sheet_names:
                if main_folder == sheet_name:
                    subfolder, label = subfolder_label.split("__")
                    sheet_dict[sheet_name]["calc_filter_list"].append({(subfolder, label, hidden): calc_filter})

        wb = Workbook()
        wb.remove(wb.active)
        for sheet_name in self.sheet_names:
            hidden_sheet = sheet_name in self.hidden_sheets
            ws = wb.create_sheet(sheet_name)
            ws.title = sheet_name
            calc_filter_list = sheet_dict[sheet_name]["calc_filter_list"]
            for identifier, (column_name, column, value_by_def) in self.columns_to_output_xlsx.items():
                ws[f"{identifier}1"] = column_name
                ws[f"{identifier}1"].font = Font(size=13, bold=True)
            row = 2
            subfolder_set = set()

            rows_to_remove = []
            for calc_filter in calc_filter_list:
                for (subfolder, label, hidden), values in calc_filter.items():
                    if len(values) > 0:
                        if subfolder == "None":
                            if row != 2 and label not in subfolder_set:
                                row = row + 1
                            ws[f"A{row}"] = label
                            if label not in subfolder_set:
                                if hidden or hidden_sheet:
                                    ws[f"A{row}"].font = Font(bold=True, color="808080")
                                else:
                                    ws[f"A{row}"].font = Font(bold=True)
                            else:
                                if hidden or hidden_sheet:
                                    ws[f"A{row}"].font = Font(color="808080")
                        else:
                            if subfolder not in subfolder_set:
                                if row != 2:
                                    row = row + 1
                                ws[f"A{row}"] = subfolder
                                if hidden or hidden_sheet:
                                    ws[f"A{row}"].font = Font(bold=True, color="808080")
                                else:
                                    ws[f"A{row}"].font = Font(bold=True)
                                row = row + 1
                                subfolder_set.add(subfolder)
                            ws[f"A{row}"] = label
                            if hidden or hidden_sheet:
                                ws[f"A{row}"].font = Font(color="808080")

                        for elem_info in values:
                            for column, value in elem_info.items():
                                ws[f"{column}{row}"] = value
                                if hidden or hidden_sheet:
                                    ws[f"{column}{row}"].font = Font(color="808080")
                            row = row + 1
                        row = row + 1
            if self.delete_hidden:
                column_names = []
                for identifier, (column_name, column, value_by_def) in self.columns_to_output_xlsx.items():
                    if column == self.hidden_field_name:
                        column_names.append(column_name)
                for column_name in column_names:
                    col = JsonHelper.search_value_in_row_index(ws, column_name)
                    ws.delete_cols(col)
            dims = {}
            for row in ws.rows:
                for cell in row:
                    alignment_obj = cell.alignment.copy(horizontal='left')
                    cell.alignment = alignment_obj
                    if cell.value:
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
            for col, value in dims.items():
                ws.column_dimensions[col].width = value
        if self.delete_hidden:
            for sheet in self.hidden_sheets:
                wb.remove(wb.get_sheet_by_name(sheet))

        wb.save(result_excel_file_name)

    @staticmethod
    def search_value_in_row_index(sheet, column_title, row=1):
        for cell in sheet[row]:
            if cell.value == column_title:
                return cell.column
        return None

    def get_folder_to_calc_and_filter(self,
                                      fields_list_with_value_by_def=None, ):
        if not fields_list_with_value_by_def:
            fields_list_with_value_by_def = self.fields_to_select_from_calc_filter

        fields_list = [field for (field, value_by_def) in fields_list_with_value_by_def]

        calculations_filter_list = []

        calculations_filter_list.extend([JsonHelper.select_fields(
            fields=fields_list,
            dictionary=calculation
        ) for calculation in self.json[self.root_calc_element_name]])

        calculations_filter_list.extend([JsonHelper.select_fields(
            fields=fields_list,
            dictionary=filter_dict
        ) for filter_dict in self.json[self.root_filter_element_name]])

        calculations_filter_dict = {
            calc_filter_info[self.identifier_field_name]:
                [calc_filter_info.get(field_name, value_by_def)
                 for (field_name, value_by_def) in fields_list_with_value_by_def
                 if field_name != self.identifier_field_name]
            for calc_filter_info in calculations_filter_list
        }

        self.get_folder_hierarchy(
            folder_list=self.json[self.root_metadata_element_name][0][self.root_folder_element_name])

        columns_to_output = list(self.columns_to_output_xlsx.keys())[1:]

        for folder, (refs, hidden) in self.folder_dict.items():
            self.folder_with_calc_filter.update({
                (folder, hidden): [{
                    col_identifier: calculations_filter_dict[ref][num]
                    for num, col_identifier in enumerate(columns_to_output)
                } for ref in refs]
            })

        if self.delete_hidden:
            folder_with_calc_filter_to_stay = {}
            folders_to_remove = [(folder, hidden) for (folder, hidden) in self.folder_with_calc_filter.keys() if hidden]
            self.folder_with_calc_filter = {key: self.folder_with_calc_filter[key]
                                            for key in self.folder_with_calc_filter if key not in folders_to_remove}
            for (folder, hidden), refs in self.folder_with_calc_filter.items():
                items_to_stay = []
                for ref in refs:
                    for column, value in ref.items():
                        if column in self.hidden_column_xlsx_list:
                            if value not in (True, "TRUE", 1):
                                items_to_stay.append(ref)
                folder_with_calc_filter_to_stay.update({(folder, hidden): items_to_stay})
            self.folder_with_calc_filter = folder_with_calc_filter_to_stay

    @staticmethod
    def select_fields(fields, dictionary):
        output_dictionary = {}
        for field in fields:
            for key, value in dictionary.items():
                if dictionary.get(field, None):
                    output_dictionary[field] = dictionary[field]
        return output_dictionary

    def get_folder_hierarchy(self, folder_list):

        def get_refs(folder_dict,
                     folder_name,
                     subfolder_name=None):

            folders_list = folder_dict[self.folder_element_name][self.root_folder_element_name]
            folder_label = folder_dict[self.folder_element_name][self.label_element_name]
            if self.hidden_field_name:
                hidden = folder_dict[self.folder_element_name].get(self.hidden_field_name, False)
            refs = []
            for element in folders_list:
                if element.get(self.ref_element_name) is not None:
                    if self.hidden_field_name:
                        hidden = element.get(self.hidden_field_name, False)
                    refs.append(element[self.ref_element_name])
                elif element.get(self.folder_element_name) is not None:
                    get_refs(element, folder_name, folder_label)
            self.sheet_names.add(folder_name)
            if folder_name == subfolder_name:
                pass
            else:
                if self.hidden_field_name:
                    self.folder_dict.update({f"{folder_name}___{subfolder_name}__{folder_label}": (refs, hidden)})
                else:
                    self.folder_dict.update({f"{folder_name}___{subfolder_name}__{folder_label}": (refs, None)})

        for folder in folder_list:
            label = folder[self.folder_element_name][self.label_element_name]
            if self.hidden_field_name:
                if folder[self.folder_element_name].get(self.hidden_field_name, False):
                    self.hidden_sheets.append(label)

            for subfolder in folder[self.folder_element_name][self.root_folder_element_name]:
                if label in self.domains_to_exclude:
                    pass
                else:
                    if subfolder.get(self.ref_element_name):
                        subfolder = {self.folder_element_name: folder[self.folder_element_name]}
                    get_refs(
                        folder_dict=subfolder,
                        folder_name=label)
